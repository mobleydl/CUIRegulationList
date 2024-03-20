import os 
import re #regex searching
from openpyxl import Workbook, load_workbook #pip install openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import fitz #pip install PyMuPDF

#tool for building accurate regex for pattern searching: https://pythex.org/

def build_file_list(files_in_dir):
    file_list = [] #empty list to store files found
    ext_pattern = r".pdf|.txt" #search for files with these extensions
    for file in files_in_dir: #iterate through files
        try:
            match = re.search(ext_pattern,file) #if .pdf or .txt add to file_lsit
            file_list.append(match.string)
        except AttributeError: #if no match, skip file
            pass
    return file_list

def build_registry():
    try: #execute if CUI Registry.xlsx file found
        reg_file = load_workbook(filename = 'CUI Registry.xlsx')
        reg_data = reg_file['Registry'] #use 'Registry' worksheet from CUI Registry.xlsx
        temp_list = [] #temp list to store Regulation,CUI Category tuples
        for row in reg_data.iter_rows(max_col=2, values_only=True):
            temp_list.append(row)
        reg_dict = dict(temp_list) #convert temp list to dictionary
        return reg_dict
    except FileNotFoundError:
        print(f'\nERROR: CUI Registry.xlsx file not found in current folder!\nVerify CUI Registry.xlsx file exists in the same folder as the RLBuilder.py script and try again!\n')
        quit()

def gen_excel_file(file_list,reg_dict):
    try:
        regulation_pattern = r"\d* USC [\d.()\-\w]* [\d.()\-\w]* [\d.()\-\w]*|\d* USC [\d.()\-\w]* [\d.()\-\w]*|\d* USC [\d.()\-\w]*|\d* CFR [\d.()\-\w,]* [\d.()\-\w,]* [\d.()\-\w,]* [\d.()\-\w]* [\d.()\-\w]*|\d* CFR [\d.()\-\w,]* [\d.()\-\w]* [\d.()\-\w]* [\d.()\-\w]*|\d* CFR [\d.()\-\w,]* [\d.()\-\w]* [\d.()\-\w]*|\d* CFR [\d.()\-\w]* [\d.()\-\w]*|\d* CFR [\d.()\-\w]*|FAR [\d.()\-\w]*|FAR [\d.()\-\w]* [\d.()\-\w]*|DFARS [\d.()\-\w]*|DFARS [\d.()\-\w]* [\d.()\-\w]*|\d* FR [\d.()\-\w]* [\d.()\-\w]* [\d.()\-\w]*|NATO [\d.()\-\w]*|Intelligence Community Directive \d*|Federal Rules of \w* \w* [\d.()\-\w]*|\d* Federal Register \d*|Executive Order \d*|HSPD \d*|NSPM \d*|FinCen BSA|GSA PBS \w* [\d.]*|P.L. [\d\-]*|OMB \w* [\w\-]* [\w]* [\d.]*|OMB \w* [\w\-]*|OMB [\w\-]*|IRS Publication \d*|United States Constitution, Article \d, Section \d|Provisional Approval [\d\-]*|Federal Continuity Directive \d*|Presidential Policy Directive \d*|United States Security Authority for NATO, Instruction [\d\-]*|The Risk Management Process for Federal Facilities: An Interagency Security Committee Standard|US Attorney's Manual"
        ext_pattern = r".pdf|.txt"
        wb = Workbook()
        for file in file_list:
            file_name = str(file)
            file_type_search = re.search(ext_pattern,file)
            file_type = file_type_search.group()

            #create new worksheet using file_name as title
            ws = wb.create_sheet(file_name.split(".")[0])

            #style header row for each newly created worksheet
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 65
            align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            font = Font(color='00FFFFFF',bold=True) #make font white
            fill = PatternFill(fill_type = 'solid',fgColor = '00003366') #make background navy blue
            A1,B1 = ws['A1'],ws['B1']
            A1.value,B1.value = 'Regulation','CUI Category'
            A1.alignment,B1.alignment = align,align
            A1.font,B1.font = font,font
            A1.fill,B1.fill = fill,fill
            ws.freeze_panes = 'A2' #freeze header panes

            if file_type == '.txt': #if txt
                file = open(file, 'r') #open file and read only
                content = file.read() #content will hold all text from .txt file
                matches = re.findall(regulation_pattern, content) #search content for regulation matches
                file.close() #close file when finished

            elif file_type == '.pdf': #if pdf
                file = fitz.open(file) #open file
                content = '' #initialize content string
                for page in file: #iterate through each pdf file page
                    content += page.get_text() #add extracted text to content string
                matches = re.findall(regulation_pattern, content) #search content for regulation matches
                file.close() #close file when finished

            #add rows to created worksheet
            for match in matches:
                match = match.rstrip() #remove trailing whitespaces that sometimes occur from reading from pdf files
                if match not in reg_dict.keys():
                    temp_row = (match,'CONFIRM REGULATION IS NOT A MATCH')
                    ws.append(temp_row)
                else:
                    for reg in reg_dict:
                        if reg == match:
                            temp_row = (match,reg_dict[reg])
                            ws.append(temp_row)
            
            #cell styling variables to be used for added rows
            columns = zip(ws['A'],ws['B']) #combine list of values from Regulation column and CUI Category column
            good_font = Font(color='006100',bold=True)
            bad_font = Font(color='9C0005',bold=True)
            c_align = Alignment(horizontal='center',vertical='center',wrap_text=True)
            r_align = Alignment(horizontal='left',vertical='center',wrap_text=True)
            good_sides = Side(border_style='thick', color='006100')
            good_border = Border(top=good_sides, left=good_sides, right = good_sides, bottom = good_sides)
            bad_sides = Side(border_style='thick', color='9C0005')
            bad_border = Border(top=bad_sides, left=bad_sides, right = bad_sides, bottom = bad_sides)

            #iterate through added rows to apply proper styling to cells based on their values
            for cellA,cellB in columns:
                if cellA.value == 'Regulation' and cellB.value == 'CUI Category':
                    pass
                elif cellB.value == 'CONFIRM REGULATION IS NOT A MATCH':
                    cellA.style,cellB.style = 'Bad','Bad'
                    cellA.font,cellB.font = bad_font,bad_font
                    cellA.alignment,cellB.alignment = r_align,c_align
                    cellA.border,cellB.border = bad_border,bad_border
                else:
                    cellA.style,cellB.style = 'Good','Good'
                    cellA.font,cellB.font = good_font,good_font
                    cellA.alignment,cellB.alignment = r_align,c_align
                    cellA.border,cellB.border = good_border,good_border

        del wb['Sheet'] #remove the default 'Sheet' worksheet that is created upon creation of new workbook
        wb.save(filename='CUI Regulation List.xlsx') #save in current directory as a new file called CUI Regulation List
        print(f'\nCUI Regulation List.xlsx created successfully!\n')

    #if user tries running script while the CUI Regulation List is open print the following:
    except PermissionError:
        print(f'\nERROR: Please close CUI Regulation List.xlsx file and try again!\n')

def main():
    files_in_dir = os.listdir(os.getcwd())
    #build list of files ending in .pdf or .txt
    file_list = build_file_list(files_in_dir)

    #check to ensure that usable contract files were found
    num_files = len(file_list)
    if num_files == 0:
        print(f'\nERROR: No contracts found in current folder!')
        print(f'Ensure that contracts are saved in the same folder as the RLBuilder.py script and the CUI Registry.xlsx file and try again!')
        print(f'WARNING: Contracts must have either a .pdf or .txt file extension to work with this script!\n')
        quit()

    #build dictionary from CUI Registry.xlsx file
    reg_dict = build_registry()

    #create excel file
    gen_excel_file(file_list,reg_dict) 

if __name__=="__main__":
    main()

#CUI Regulation List Builder script
#Version 3.0
#Created by Dustin Mobley
#Updated: 3/20/2024